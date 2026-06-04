import os
import logging
from typing import Tuple, List
from app.services.com_wrapper import get_excel_app, HAS_COM
from app.utils.paths import resolve_dynamic_path

logger = logging.getLogger("DocBuilder.ExcelValidator")

# Optional openpyxl for macOS/Linux validation
HAS_OPENPYXL = False
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    pass

def validate_excel_item(excel_path: str, sheet_name: str, range_address: str = None, chart_id: int = None, config_path: str = "") -> Tuple[bool, List[str]]:
    """
    Validates Excel configuration.
    Returns (is_valid, errors_list)
    """
    errors = []
    
    # 1. Path resolution and existence check
    if not excel_path:
        return False, ["Excel file path is empty."]
        
    resolved_path = resolve_dynamic_path(excel_path, config_path)
    if not os.path.exists(resolved_path):
        return False, [f"Excel file does not exist: {excel_path} (Resolved local: {resolved_path})"]

    # 2. Validation strategy
    if HAS_COM:
        # Windows validation using Excel COM
        excel = None
        wb = None
        try:
            excel = get_excel_app()
            excel.Visible = False
            excel.DisplayAlerts = False
            
            wb = excel.Workbooks.Open(resolved_path, ReadOnly=True)
            
            # Check sheet
            sheet_found = False
            ws = None
            try:
                ws = wb.Worksheets(sheet_name)
                sheet_found = True
            except Exception:
                errors.append(f"Sheet '{sheet_name}' not found in workbook.")
                
            if sheet_found and ws:
                # Validate range if provided
                if range_address:
                    try:
                        rng = ws.Range(range_address)
                        # Access an attribute to confirm it is valid
                        _ = rng.Address
                    except Exception:
                        errors.append(f"Invalid range address: '{range_address}'")
                
                # Validate chart_id if provided
                if chart_id is not None:
                    try:
                        # ChartObjects in VBA/COM are 1-indexed. Let's verify if index exists
                        count = ws.ChartObjects().Count
                        if count == 0:
                            errors.append(f"No charts found on sheet '{sheet_name}'.")
                        elif chart_id < 1 or chart_id > count:
                            errors.append(f"Chart ID {chart_id} out of bounds. Found {count} charts on sheet.")
                    except Exception as ce:
                        errors.append(f"Failed to check charts: {ce}")
                        
        except Exception as e:
            errors.append(f"Failed to open workbook via Excel: {e}")
        finally:
            if wb:
                try:
                    wb.Close(False)
                except Exception:
                    pass
            if excel:
                try:
                    excel.Quit()
                except Exception:
                    pass
    else:
        # Cross-platform validation (macOS / Linux)
        if HAS_OPENPYXL:
            try:
                # Open in read-only and data_only mode (fast)
                wb = openpyxl.load_workbook(resolved_path, read_only=True, data_only=True)
                if sheet_name not in wb.sheetnames:
                    errors.append(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}")
                else:
                    if range_address:
                        from openpyxl.utils.cell import coordinate_to_tuple
                        if ":" not in range_address:
                            # Single cell coordinate check
                            try:
                                coordinate_to_tuple(range_address)
                            except Exception:
                                errors.append(f"Range address '{range_address}' is not a valid cell coordinate (e.g. A1 or G34).")
                        else:
                            try:
                                # Try parsing range coordinate
                                parts = range_address.split(":")
                                for part in parts:
                                    coordinate_to_tuple(part)
                            except Exception:
                                errors.append(f"Invalid cell coordinates in range: '{range_address}'")
                    
                    if chart_id is not None:
                        # openpyxl can read drawings, but chart_id mapping might be different.
                        # Let's just log a warning that chart counts cannot be fully validated on Mac.
                        ws = wb[sheet_name]
                        charts_count = len(getattr(ws, "_charts", []))
                        logger.info(f"Detected {charts_count} charts on sheet '{sheet_name}' (via openpyxl).")
                        if charts_count == 0:
                            logger.warning("No openpyxl charts detected, but they might be legacy/COM charts. Skipping strict check on Mac.")
            except Exception as e:
                errors.append(f"Failed to parse Excel file via openpyxl: {e}")
        else:
            errors.append("openpyxl is not installed. Sheet and range validation skipped on macOS.")

    is_valid = len(errors) == 0
    return is_valid, errors
